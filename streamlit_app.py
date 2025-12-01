import streamlit as st
import pandas as pd
import os

EXCEL_FILE = "account_planning_tool.xlsx"
MASTER_SHEET = "Master_overview"

st.set_page_config(page_title="Account Planning Input", layout="wide")

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=30)
def load_master(path: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=MASTER_SHEET, engine="openpyxl")

def save_master(df: pd.DataFrame, path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb[MASTER_SHEET]
    # overwrite all rows/cols of the original sheet with df
    for r_idx, row in enumerate(df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(path)

def build_index(df: pd.DataFrame):
    """
    The Master_overview is structured in 5-row blocks per account:
    row 0: Rep, Account, ..., Reason, ..., Current State, ...
    row 1: Score, Description, ..., Ideal State
    row 2: Score (Gap),..., Gap
    row 3: Score (Action),..., Action
    row 4: Score (Review),..., Review
    """
    blocks = []
    i = 0
    while i + 4 < len(df):
        rep = df.iloc[i, 0]
        account = df.iloc[i, 1]
        reason = df.iloc[i, 4]
        if pd.notna(rep) and pd.notna(account):
            blocks.append(
                {
                    "start_row": i,
                    "rep": str(rep),
                    "account": str(account),
                    "reason": str(reason) if pd.notna(reason) else "",
                }
            )
        i += 5
    index_df = pd.DataFrame(blocks)
    return index_df

def load_block(df: pd.DataFrame, start_row: int):
    """
    Read one account block starting at start_row from Master_overview.
    Adjust column indices to match your exact layout if needed.
    """
    r0 = start_row
    r1 = start_row + 1
    r2 = start_row + 2
    r3 = start_row + 3
    r4 = start_row + 4

    rep = df.iloc[r0, 0]
    account = df.iloc[r0, 1]
    reason = df.iloc[r0, 4]

    current_state = df.iloc[r0, 7] if df.shape[1] > 7 else ""
    ideal_state   = df.iloc[r1, 7] if df.shape[1] > 7 else ""
    gap           = df.iloc[r2, 7] if df.shape[1] > 7 else ""
    action        = df.iloc[r3, 7] if df.shape[1] > 7 else ""
    review        = df.iloc[r4, 7] if df.shape[1] > 7 else ""

    score = df.iloc[r1, 0]  # the main score row

    return {
        "rep": rep,
        "account": account,
        "reason": reason,
        "score": score,
        "current_state": current_state if pd.notna(current_state) else "",
        "ideal_state": ideal_state if pd.notna(ideal_state) else "",
        "gap": gap if pd.notna(gap) else "",
        "action": action if pd.notna(action) else "",
        "review": review if pd.notna(review) else "",
    }

def write_block(df: pd.DataFrame, start_row: int, data: dict) -> pd.DataFrame:
    """
    Write edited input back into the Master_overview DataFrame.
    """
    r0 = start_row
    r1 = start_row + 1
    r2 = start_row + 2
    r3 = start_row + 3
    r4 = start_row + 4

    df.iloc[r1, 0] = data["score"]              # main score
    if df.shape[1] > 7:
        df.iloc[r0, 7] = data["current_state"]
        df.iloc[r1, 7] = data["ideal_state"]
        df.iloc[r2, 7] = data["gap"]
        df.iloc[r3, 7] = data["action"]
        df.iloc[r4, 7] = data["review"]
    return df

# ─────────────────────────────────────────────────────────────
# Main app
# ─────────────────────────────────────────────────────────────
def main():
    st.title("Account Planning – Data Entry")

    if not os.path.exists(EXCEL_FILE):
        st.error(f"Excel file '{EXCEL_FILE}' not found. Please upload it.")
        uploaded = st.file_uploader("Upload account_planning_tool.xlsx", type=["xlsx"])
        if uploaded:
            with open(EXCEL_FILE, "wb") as f:
                f.write(uploaded.getbuffer())
            st.success("File uploaded. Please reload the page.")
        return

    master_df = load_master(EXCEL_FILE)
    index_df = build_index(master_df)

    if index_df.empty:
        st.warning("No rep/account blocks detected in Master_overview.")
        st.dataframe(master_df)
        return

    # Dropdowns
    reps = sorted(index_df["rep"].unique())
    selected_rep = st.selectbox("Select Sales Rep", reps)

    accounts = sorted(index_df[index_df["rep"] == selected_rep]["account"].unique())
    selected_account = st.selectbox("Select Account", accounts)

    # Find the block for this rep + account
    row = index_df[(index_df["rep"] == selected_rep) &
                   (index_df["account"] == selected_account)].iloc[0]
    start_row = int(row["start_row"])

    block = load_block(master_df, start_row)

    st.subheader(f"{selected_rep} – {selected_account}")
    st.caption(f"Reason: {block['reason']}")

    col1, col2 = st.columns([1, 3])
    with col1:
        score = st.number_input("Overall Score", value=int(block["score"]), step=1)

    with col2:
        current_state = st.text_area("Current State", value=block["current_state"], height=80)
        ideal_state   = st.text_area("Ideal State",   value=block["ideal_state"],   height=80)
        gap           = st.text_area("Gap",           value=block["gap"],           height=80)
        action        = st.text_area("Action",        value=block["action"],        height=80)
        review        = st.text_area("Review",        value=block["review"],        height=80)

    if st.button("💾 Save changes"):
        edited = {
            "score": score,
            "current_state": current_state,
            "ideal_state": ideal_state,
            "gap": gap,
            "action": action,
            "review": review,
        }
        updated_df = write_block(master_df.copy(), start_row, edited)
        save_master(updated_df, EXCEL_FILE)
        load_master.clear()  # clear cache
        st.success("Changes saved to Excel.")

    with st.expander("Show raw Master_overview (read-only)"):
        st.dataframe(master_df, use_container_width=True)

if __name__ == "__main__":
    main()
